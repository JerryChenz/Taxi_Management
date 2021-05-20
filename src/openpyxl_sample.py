'''
from tkinter import *

root = Tk()

var = StringVar()
# Creating a Label Widget
label = Label( root, textvariable=var, relief=RAISED )

var.set("Hey!? How are you doing?")
# Shoving it onto the screen
label.pack()
root.mainloop()
'''

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
wb = Workbook()

# load existing spreadsheet
wb = load_workbook('Z:\Taxi Management System\M1 - Drivers Management/Rent Schedule_with v0.3.xlsm')

# Create a active worksheet
ws = wb.active

# Set Variables
driver_id = ws["A2"].value
name_cn = ws["B2"].value
column_a = ws['A']
column_b = ws['B']
row_1 = ws['1']
print (column_a)

# print something from our spreadsheet
print(ws["A2"].value)
print (f'{driver_id}:{name_cn}')

for cell in column_a:
    print(f'{cell.value}\n')

for cell in row_1:
    print(f'{cell.value}\n')